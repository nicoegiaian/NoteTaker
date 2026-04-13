import sys
from pathlib import Path

def leer_word(ruta):
    from docx import Document
    doc = Document(ruta)
    texto = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
    return texto[:500]  # primeros 500 caracteres

def leer_pdf(ruta):
    import pdfplumber
    with pdfplumber.open(ruta) as pdf:
        texto = ""
        for page in pdf.pages[:2]:  # primeras 2 páginas
            texto += page.extract_text() or ""
    return texto[:500]

def leer_excel(ruta):
    from openpyxl import load_workbook
    wb = load_workbook(ruta, data_only=True)
    ws = wb.active
    filas = []
    for row in list(ws.iter_rows(values_only=True))[:10]:  # primeras 10 filas
        fila = [str(c) for c in row if c is not None]
        if fila:
            filas.append(" | ".join(fila))
    return "\n".join(filas)

# ─── CONFIGURÁ ESTAS RUTAS ───────────────────────────────
RUTA_WORD  = r"C:\Users\degiaian\OneDrive - ASE Conecta\CORPO - Gerencia de Proyectos Corporativos - Proyectos en curso\DevSecOps\6- Adopción\Apigee\Requisitos_OpenAPI_Apigee.docx"
RUTA_PDF   = r"C:\Users\degiaian\OneDrive - ASE Conecta\CORPO - Gerencia de Proyectos Corporativos - Proyectos en curso\DevSecOps\6- Adopción\ESTRATEGIA-ARQUITECTURA_DE_PLATAFORMA-2026.pdf"
RUTA_EXCEL = r"C:\Users\degiaian\OneDrive - ASE Conecta\CORPO - Gerencia de Proyectos Corporativos - Proyectos en curso\DevSecOps\6- Adopción\Avance adopción celulas.xlsx"
# ─────────────────────────────────────────────────────────

print("=" * 50)
print("POC — Lectura de archivos de SharePoint local")
print("=" * 50)

print("\n📄 WORD:")
try:
    print(leer_word(RUTA_WORD))
except Exception as e:
    print(f"Error: {e}")

print("\n📕 PDF:")
try:
    print(leer_pdf(RUTA_PDF))
except Exception as e:
    print(f"Error: {e}")

print("\n📊 EXCEL:")
try:
    print(leer_excel(RUTA_EXCEL))
except Exception as e:
    print(f"Error: {e}")

print("\n✅ POC completada")