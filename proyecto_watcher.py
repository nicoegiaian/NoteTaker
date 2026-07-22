"""
proyecto_watcher.py — F1b: Monitoreo y digest diario de archivos de proyectos
"""

import os
import json
import logging
import requests
import urllib3
from pathlib import Path
from datetime import datetime, date, timedelta

# Toda la configuración tuneable vive en configuracion.toml (leída por config.py).
from config import (
    PROYECTOS as PROYECTOS_KEYWORDS,        # palabra clave → proyecto
    PROYECTOS_POR_CARPETA as PROYECTOS,     # carpeta → proyecto (F1b)
    EXTENSIONES_SOPORTADAS,
    CONTEXTO_FOLDER, NOVEDADES_FOLDER, MAILS_FOLDER, COLA_FILE, DIGEST_LOG_FILE,
    MODELO_DEFAULT, MODELO_DIGEST, MATRIZ_PATHS,
    PRECIOS_USD as _PRECIOS_USD,
    ASUNTOS_RUIDO as _ASUNTOS_RUIDO,
    MARCADORES_CITA as _MARCADORES_CITA,
    DIGESTS_HISTORICOS,
    PROMPTS_FOLDER,
)

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
log = logging.getLogger(__name__)

# Endpoint de la API (constante, no es configuración de usuario).
ANTHROPIC_API_URL = "https://api.anthropic.com/v1/messages"


# ─── COLA ─────────────────────────────────────────────────
def cargar_cola() -> dict:
    if Path(COLA_FILE).exists():
        try:
            return json.loads(Path(COLA_FILE).read_text(encoding="utf-8"))
        except Exception:
            pass
    return {nombre: [] for nombre in PROYECTOS.values()}


def guardar_cola(cola: dict):
    Path(COLA_FILE).write_text(
        json.dumps(cola, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def encolar_archivo(ruta: str, evento: str):
    """Agrega un archivo a la cola del proyecto correspondiente."""
    proyecto = detectar_proyecto(ruta)
    if not proyecto:
        return
    if Path(ruta).suffix.lower() not in EXTENSIONES_SOPORTADAS:
        return
    if Path(ruta).name.startswith("~$"):
        return

    cola = cargar_cola()
    if proyecto not in cola:
        cola[proyecto] = []

    # Evitar duplicados de la misma ruta
    rutas_existentes = [item["ruta"] for item in cola[proyecto]]
    if ruta in rutas_existentes:
        # Actualizar evento si ya existe
        for item in cola[proyecto]:
            if item["ruta"] == ruta:
                item["evento"] = evento
                item["timestamp"] = datetime.now().isoformat()
        log.info(f"[Cola] Actualizado: {Path(ruta).name} → {proyecto}")
    else:
        cola[proyecto].append({
            "ruta": ruta,
            "nombre": Path(ruta).name,
            "evento": evento,
            "timestamp": datetime.now().isoformat(),
        })
        log.info(f"[Cola] Encolado: {Path(ruta).name} → {proyecto}")

    guardar_cola(cola)


# ─── DIGEST LOG ───────────────────────────────────────────
def cargar_digest_log() -> dict:
    if Path(DIGEST_LOG_FILE).exists():
        try:
            return json.loads(Path(DIGEST_LOG_FILE).read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def guardar_digest_log(log_data: dict):
    Path(DIGEST_LOG_FILE).write_text(
        json.dumps(log_data, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def proyectos_procesados_hoy() -> set:
    """Retorna el set de proyectos ya procesados hoy."""
    log_data = cargar_digest_log()
    ultima = log_data.get("fecha", "")
    try:
        if datetime.fromisoformat(ultima).date() == date.today():
            return set(log_data.get("proyectos_ok", []))
    except Exception:
        pass
    return set()


def marcar_proyecto_procesado(proyecto: str):
    """Marca un proyecto como procesado exitosamente hoy."""
    log_data = cargar_digest_log()
    fecha_actual = date.today().isoformat()

    # Si cambió el día, resetear
    try:
        if log_data.get("fecha", "") != fecha_actual:
            log_data = {"fecha": fecha_actual, "proyectos_ok": []}
    except Exception:
        log_data = {"fecha": fecha_actual, "proyectos_ok": []}

    if proyecto not in log_data.get("proyectos_ok", []):
        log_data.setdefault("proyectos_ok", []).append(proyecto)

    log_data["ultima_ejecucion"] = datetime.now().isoformat()
    guardar_digest_log(log_data)


def digest_ya_corrido_hoy() -> bool:
    """El digest está completo si todos los proyectos fueron procesados hoy."""
    procesados = proyectos_procesados_hoy()
    todos = set(PROYECTOS.values())
    return todos.issubset(procesados)


# ─── LECTURA DE ARCHIVOS ──────────────────────────────────
def detectar_proyecto(ruta: str) -> str | None:
    for carpeta_proyecto, nombre_proyecto in PROYECTOS.items():
        if ruta.startswith(carpeta_proyecto):
            return nombre_proyecto
    return None


def leer_word(ruta: str) -> str:
    from docx import Document
    doc = Document(ruta)
    texto = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
    return texto[:8000]


def leer_pdf(ruta: str) -> str:
    import pdfplumber
    with pdfplumber.open(ruta) as pdf:
        texto = ""
        for page in pdf.pages[:5]:
            texto += page.extract_text() or ""
    return texto[:8000]


def leer_excel(ruta: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(ruta, data_only=True)
    resultado = []
    for nombre_hoja in wb.sheetnames[:3]:
        ws = wb[nombre_hoja]
        filas = []
        for row in list(ws.iter_rows(values_only=True))[:30]:
            fila = [str(c) for c in row if c is not None]
            if fila:
                filas.append(" | ".join(fila))
        if filas:
            resultado.append(f"[Hoja: {nombre_hoja}]\n" + "\n".join(filas))
    return "\n\n".join(resultado)[:8000]


def leer_contenido(ruta: str) -> str | None:
    ext = Path(ruta).suffix.lower()
    try:
        if ext == ".docx":
            return leer_word(ruta)
        elif ext == ".pdf":
            return leer_pdf(ruta)
        elif ext in {".xlsx", ".xls"}:
            return leer_excel(ruta)
    except Exception as e:
        log.error(f"Error leyendo {ruta}: {e}")
    return None


# ─── CONTEXTO ─────────────────────────────────────────────
def cargar_contexto(proyecto: str) -> str:
    ruta = Path(CONTEXTO_FOLDER) / f"contexto_{proyecto}.txt"
    if ruta.exists():
        return ruta.read_text(encoding="utf-8")
    return "Sin contexto previo disponible."


def guardar_contexto(proyecto: str, contenido_agente: str):
    """
    Actualiza solo las secciones del agente en el contexto del proyecto.
    Nunca toca las secciones escritas por el PM:
    DESCRIPCIÓN Y ALCANCE, EQUIPO CLAVE, HITOS VIGENTES, RIESGOS IDENTIFICADOS.
    """
    ruta = Path(CONTEXTO_FOLDER) / f"contexto_{proyecto}.txt"
    
    # Secciones protegidas — escritas por el PM
    SECCIONES_PROTEGIDAS = [
        "=== DESCRIPCIÓN Y ALCANCE ===",
        "=== EQUIPO CLAVE ===",
        "=== HITOS VIGENTES ===",
        "=== RIESGOS IDENTIFICADOS ===",
    ]
    
    # Secciones del agente — se actualizan automáticamente
    SECCIONES_AGENTE = [
        "=== ESTADO ACTUAL ===",
        "=== DECISIONES TOMADAS ===",
    ]

    if ruta.exists():
        contenido_actual = ruta.read_text(encoding="utf-8")
    else:
        contenido_actual = ""

    # Verificar si el archivo tiene la estructura esperada
    tiene_estructura = any(s in contenido_actual for s in SECCIONES_PROTEGIDAS)

    if not tiene_estructura:
        # Archivo sin estructura → escribir todo (comportamiento anterior)
        ruta.write_text(contenido_agente, encoding="utf-8")
        log.info(f"Contexto creado: {ruta.name}")
        return

    # Extraer la parte protegida (todo hasta la primera sección del agente)
    primera_seccion_agente = None
    pos_corte = len(contenido_actual)
    for seccion in SECCIONES_AGENTE:
        pos = contenido_actual.find(seccion)
        if pos != -1 and pos < pos_corte:
            pos_corte = pos
            primera_seccion_agente = seccion

    parte_protegida = contenido_actual[:pos_corte].rstrip()

    # Parsear el contenido nuevo del agente para extraer estado y decisiones
    estado_nuevo = ""
    decisiones_nuevas = ""

    if "=== ESTADO ACTUAL ===" in contenido_agente:
        partes = contenido_agente.split("=== ESTADO ACTUAL ===")
        if len(partes) > 1:
            resto = partes[1]
            if "=== DECISIONES TOMADAS ===" in resto:
                estado_nuevo = resto.split("=== DECISIONES TOMADAS ===")[0].strip()
            else:
                estado_nuevo = resto.strip()

    if "=== DECISIONES TOMADAS ===" in contenido_agente:
        partes = contenido_agente.split("=== DECISIONES TOMADAS ===")
        if len(partes) > 1:
            decisiones_nuevas = partes[1].strip()

    # Si el agente no devolvió secciones estructuradas, preservar las existentes
    if not estado_nuevo:
        if "=== ESTADO ACTUAL ===" in contenido_actual:
            partes = contenido_actual.split("=== ESTADO ACTUAL ===")
            resto = partes[1] if len(partes) > 1 else ""
            if "=== DECISIONES TOMADAS ===" in resto:
                estado_nuevo = resto.split("=== DECISIONES TOMADAS ===")[0].strip()
            else:
                estado_nuevo = resto.strip()

    if not decisiones_nuevas:
        if "=== DECISIONES TOMADAS ===" in contenido_actual:
            partes = contenido_actual.split("=== DECISIONES TOMADAS ===")
            decisiones_nuevas = partes[1].strip() if len(partes) > 1 else ""

    # Reconstruir el archivo completo
    nuevo_contenido = f"""{parte_protegida}

=== ESTADO ACTUAL ===
{estado_nuevo}

=== DECISIONES TOMADAS ===
{decisiones_nuevas}
"""
    ruta.write_text(nuevo_contenido, encoding="utf-8")
    log.info(f"Contexto actualizado (secciones protegidas intactas): {ruta.name}")



# ─── CLAUDE API ───────────────────────────────────────────
def llamar_claude(prompt: str, modelo: str = MODELO_DEFAULT, sin_pensar: bool = False) -> dict:
    api_key = os.getenv("ANTHROPIC_API_KEY")
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    payload = {
        "model": modelo,
        "max_tokens": 2500,
        "messages": [{"role": "user", "content": prompt}],
    }
    # Sonnet 5 trae thinking adaptativo por defecto; lo desactivamos para el digest
    # (salida y costo predecibles, sin que el thinking consuma el presupuesto de tokens).
    if sin_pensar:
        payload["thinking"] = {"type": "disabled"}

    resp = requests.post(
        ANTHROPIC_API_URL,
        headers=headers,
        json=payload,
        verify=False,
        timeout=90,
    )
    resp.raise_for_status()
    data = resp.json()

    # Extraer el primer bloque de texto (puede haber bloques 'thinking' antes)
    texto = next((b.get("text", "") for b in data.get("content", []) if b.get("type") == "text"), "")

    return {
        "texto": texto,
        "input_tokens": data["usage"]["input_tokens"],
        "output_tokens": data["usage"]["output_tokens"],
    }


# ─── TOKENS ───────────────────────────────────────────────
def registrar_tokens(proyecto: str, funcion: str, input_tokens: int, output_tokens: int,
                     modelo: str = MODELO_DEFAULT):
    try:
        from openpyxl import load_workbook
        ruta = Path(CONTEXTO_FOLDER) / "PM_Agent_Registro.xlsx"
        wb = load_workbook(ruta)
        ws = wb.active
        precio_in, precio_out = _PRECIOS_USD.get(modelo, _PRECIOS_USD[MODELO_DEFAULT])
        costo = (input_tokens * precio_in) + (output_tokens * precio_out)
        ws.append([
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            proyecto,
            funcion,
            input_tokens,
            output_tokens,
            round(costo, 6),
        ])
        wb.save(ruta)
    except Exception as e:
        log.warning(f"No se pudo registrar tokens: {e}")


# ─── NOVEDADES ────────────────────────────────────────────
def escribir_novedad(proyecto: str, mensaje: str, sufijo: str):
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre = f"{proyecto}_{timestamp}_{sufijo}.txt"
        ruta = Path(NOVEDADES_FOLDER) / nombre
        ruta.write_text(mensaje, encoding="utf-8")
        log.info(f"   Novedad escrita: {nombre}")
    except Exception as e:
        log.error(f"Error escribiendo novedad: {e}")


# ─── FUENTES DEL DIGEST DIARIO ────────────────────────────
def _detectar_proyecto_por_nombre(nombre: str) -> str | None:
    """Detecta el proyecto por palabra clave en el nombre de archivo (minutas)."""
    nombre_lower = nombre.lower()
    for clave, proyecto in PROYECTOS_KEYWORDS.items():
        if clave.lower() in nombre_lower:
            return proyecto
    return None


def _es_ruido(subject: str) -> bool:
    s = (subject or "").strip().lower()
    return any(marcador in s for marcador in _ASUNTOS_RUIDO)


def _cortar_cita(texto: str) -> str:
    """Corta la cita del hilo anterior: devuelve solo lo nuevo que se escribió."""
    pos_corte = len(texto)
    for marcador in _MARCADORES_CITA:
        pos = texto.find(marcador)
        if pos != -1:
            pos_corte = min(pos_corte, pos)
    return texto[:pos_corte].strip()


def _limpiar_mail(mail: dict, limite: int = 2000) -> str | None:
    """Convierte un mail (dict del JSON de Power Automate) en un bloque de texto
    limpio para el prompt, o None si es ruido / queda vacío."""
    subject = mail.get("subject", "")
    if _es_ruido(subject):
        return None

    # Cuerpo: preferir 'body' (HTML completo); caer a 'bodyPreview' si no está.
    crudo = mail.get("body") or mail.get("bodyPreview") or ""
    if "<" in crudo and ">" in crudo:
        texto = _html_str_a_texto(crudo)
    else:
        texto = crudo
    texto = _cortar_cita(texto)
    if len(texto) > limite:
        texto = texto[:limite] + "…"
    if not texto:
        return None

    cab = (f"De: {mail.get('from', '')} | "
           f"Para: {mail.get('toRecipients', '')} | "
           f"{mail.get('receivedDateTime', '')}\n"
           f"Asunto: {subject}")
    return f"{cab}\n{texto}"


def _archivos_del_proyecto(carpeta: Path, proyecto: str, fecha: str) -> list:
    """Todos los archivos de mails del día que pertenecen al proyecto: el de
    nombre exacto y cualquiera cuyo nombre contenga una palabra clave del
    proyecto. Incluye los de correos enviados, p. ej.
    'Salesforce_enviados_AAAAMMDD.txt', que se juntan con los de entrada."""
    rutas = []
    exacto = carpeta / f"{proyecto}_{fecha}.txt"
    if exacto.exists():
        rutas.append(exacto)
    for f in sorted(carpeta.glob(f"*_{fecha}.txt")):
        if f == exacto:
            continue
        if _detectar_proyecto_por_nombre(f.name) == proyecto:
            rutas.append(f)
    return rutas


def leer_mails_del_dia(proyecto: str) -> str:
    """Junta TODOS los archivos de mails que Power Automate dejó hoy para el
    proyecto (entrada + enviados) y los devuelve limpios: parsea el JSON, extrae
    el cuerpo (HTML→texto), corta la cita del hilo, descarta notificaciones de
    reunión y respuestas automáticas, y deduplica.

    Acepta el nombre exacto ('Programa Salesforce_AAAAMMDD.txt') o cualquier
    archivo del día cuyo nombre contenga una palabra clave del proyecto
    ('Salesforce_AAAAMMDD.txt', 'Salesforce_enviados_AAAAMMDD.txt')."""
    fecha = date.today().strftime("%Y%m%d")
    carpeta = Path(MAILS_FOLDER)
    if not carpeta.exists():
        return ""

    rutas = _archivos_del_proyecto(carpeta, proyecto, fecha)
    if not rutas:
        return ""

    mails = []
    crudos_no_json = []  # compatibilidad con archivos en formato viejo / texto plano
    for ruta in rutas:
        try:
            contenido = ruta.read_text(encoding="utf-8").strip()
        except Exception as e:
            log.warning(f"[Digest] No se pudo leer mails de {proyecto} ({ruta.name}): {e}")
            continue
        if not contenido or contenido == "[]":
            continue
        try:
            data = json.loads(contenido)
        except Exception:
            crudos_no_json.append(contenido)
            continue
        if isinstance(data, list):
            mails.extend(m for m in data if isinstance(m, dict))
        elif isinstance(data, dict):
            mails.append(data)

    # Limpiar + deduplicar (mismo mail capturado en más de un archivo)
    vistos = set()
    bloques = []
    for mail in mails:
        clave = (mail.get("from", ""), mail.get("subject", ""),
                 mail.get("receivedDateTime", ""))
        if clave in vistos:
            continue
        vistos.add(clave)
        bloque = _limpiar_mail(mail)
        if bloque:
            bloques.append(bloque)

    bloques.extend(crudos_no_json)
    return "\n\n---\n\n".join(bloques)


def _filas_con_encabezado(ws, marcador: str = "ID"):
    """Encuentra la fila de encabezado (la que contiene `marcador`) y devuelve
    los registros siguientes como dicts {encabezado: valor}."""
    filas = list(ws.iter_rows(values_only=True))
    headers = None
    corte = 0
    for i, fila in enumerate(filas):
        celdas = [str(c).strip() if c is not None else "" for c in fila]
        if marcador in celdas:
            headers = celdas
            corte = i + 1
            break
    if headers is None:
        return []
    registros = []
    for fila in filas[corte:]:
        vals = ["" if c is None else str(c).strip() for c in fila]
        reg = {h: (vals[j] if j < len(vals) else "") for j, h in enumerate(headers) if h}
        registros.append(reg)
    return registros


def leer_matriz(proyecto: str) -> str:
    """Lee la matriz de gobierno del proyecto (solapas de Riesgos y Decisiones)
    y la devuelve como texto para el prompt. Solo proyectos en MATRIZ_PATHS."""
    ruta = MATRIZ_PATHS.get(proyecto)
    if not ruta or not Path(ruta).exists():
        return ""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(ruta, data_only=True, read_only=True)
    except Exception as e:
        log.warning(f"[Digest] No se pudo abrir la matriz de {proyecto}: {e}")
        return ""

    def _corto(txt, n=400):
        return txt[:n] + "…" if len(txt) > n else txt

    partes = []
    try:
        # ── Riesgos ──
        if "Matriz de Riesgos" in wb.sheetnames:
            lineas = []
            for r in _filas_con_encabezado(wb["Matriz de Riesgos"]):
                rid    = r.get("ID", "")
                nombre = r.get("Nombre", "")
                desc   = r.get("Descripcion", "") or r.get("Descripción", "")
                if not rid or not (nombre or desc):
                    continue
                cab = " | ".join(x for x in [
                    rid, nombre,
                    f"Criticidad: {r.get('Criticidad','')}" if r.get("Criticidad") else "",
                    f"Estado: {r.get('Estado','')}"         if r.get("Estado") else "",
                    f"Plazo: {r.get('Plazo','')}"           if r.get("Plazo") else "",
                    f"Resp: {r.get('Responsable','')}"      if r.get("Responsable") else "",
                ] if x)
                bloque = f"- {cab}"
                if desc:
                    bloque += f"\n    Descripción: {_corto(desc)}"
                if r.get("Acciones a Tomar"):
                    bloque += f"\n    Acción: {_corto(r['Acciones a Tomar'])}"
                lineas.append(bloque)
            if lineas:
                partes.append("=== RIESGOS REGISTRADOS ===\n" + "\n".join(lineas))

        # ── Decisiones ──
        if "Decision Log" in wb.sheetnames:
            lineas = []
            for r in _filas_con_encabezado(wb["Decision Log"]):
                did = r.get("ID", "")
                det = r.get("Detalle", "")
                if not did or not det:
                    continue
                cab = " | ".join(x for x in [
                    did,
                    r.get("Fecha Inicio", ""),
                    f"Estado: {r.get('Estado','')}"   if r.get("Estado") else "",
                    f"Impacto: {r.get('Impacto','')}" if r.get("Impacto") else "",
                ] if x)
                bloque = f"- {cab}\n    {_corto(det)}"
                if r.get("Justificación"):
                    bloque += f"\n    Justificación: {_corto(r['Justificación'])}"
                if r.get("Vínculo a Riesgo / Dependencia"):
                    bloque += f"\n    Vínculo: {r['Vínculo a Riesgo / Dependencia']}"
                lineas.append(bloque)
            if lineas:
                partes.append("=== DECISIONES REGISTRADAS ===\n" + "\n".join(lineas))
    finally:
        wb.close()

    return "\n\n".join(partes)


def _html_str_a_texto(html: str, limite: int = 10000) -> str:
    """Extrae texto plano de un string HTML (cuerpo de mail o minuta)."""
    from bs4 import BeautifulSoup
    try:
        soup = BeautifulSoup(html, "html.parser")
        return soup.get_text(separator="\n", strip=True)[:limite]
    except Exception:
        return html[:limite]


def _html_a_texto(ruta: str, limite: int = 3000) -> str:
    """Extrae texto plano de una minuta HTML (archivo)."""
    try:
        contenido = Path(ruta).read_text(encoding="utf-8")
    except Exception as e:
        log.warning(f"[Digest] No se pudo leer minuta {Path(ruta).name}: {e}")
        return ""
    return _html_str_a_texto(contenido, limite)


def leer_reuniones_recientes(proyecto: str, desde: datetime, maximo: int = 5) -> str:
    """Junta el texto de las minutas (HTML) del proyecto generadas desde `desde`."""
    output = os.getenv("OUTPUT_FOLDER", "")
    if not output or not Path(output).exists():
        return ""

    partes = []
    htmls = sorted(Path(output).glob("*.html"), key=lambda f: f.stat().st_mtime, reverse=True)
    for html in htmls:
        if _detectar_proyecto_por_nombre(html.name) != proyecto:
            continue
        if datetime.fromtimestamp(html.stat().st_mtime) < desde:
            continue
        texto = _html_a_texto(str(html))
        if texto:
            fecha = datetime.fromtimestamp(html.stat().st_mtime).strftime("%d/%m/%Y")
            partes.append(f"--- REUNIÓN {fecha}: {html.name} ---\n{texto}")
        if len(partes) >= maximo:
            break

    return "\n\n".join(partes)


def _cutoff_reuniones() -> datetime:
    """Desde cuándo juntar reuniones: última ejecución del digest, o 3 días atrás."""
    log_data = cargar_digest_log()
    try:
        return datetime.fromisoformat(log_data.get("ultima_ejecucion", ""))
    except Exception:
        return datetime.now() - timedelta(days=3)


def leer_digests_recientes(proyecto: str, n: int = DIGESTS_HISTORICOS) -> str:
    """Junta los últimos `n` digests diarios del proyecto (de Novedades), del más
    viejo al más nuevo, para que el razonamiento detecte tendencias y
    estancamientos. Excluye el digest de hoy (aún no escrito, o de un reintento).
    Es solo contexto histórico: NO dispara el digest por sí mismo."""
    carpeta = Path(NOVEDADES_FOLDER)
    if n <= 0 or not carpeta.exists():
        return ""

    sufijo = "_digest_diario.txt"
    hoy = date.today().strftime("%Y%m%d")
    prefijo = f"{proyecto}_"

    # El timestamp va en el nombre (AAAAMMDD_HHMMSS), así que el orden alfabético
    # equivale al cronológico.
    archivos = sorted(carpeta.glob(f"{prefijo}*{sufijo}"))
    archivos = [f for f in archivos if f.name[len(prefijo):len(prefijo) + 8] != hoy]
    archivos = archivos[-n:]  # los n más recientes, en orden cronológico

    partes = []
    for f in archivos:
        try:
            texto = f.read_text(encoding="utf-8").strip()
        except Exception:
            continue
        if not texto:
            continue
        fecha_str = f.name[len(prefijo):len(prefijo) + 8]
        try:
            fecha = datetime.strptime(fecha_str, "%Y%m%d").strftime("%d/%m/%Y")
        except Exception:
            fecha = fecha_str
        partes.append(f"--- DIGEST {fecha} ---\n{texto}")

    return "\n\n".join(partes)


# ─── DIGEST ───────────────────────────────────────────────
def digest_proyecto(proyecto: str, items: list, cutoff_reuniones: datetime):
    """Genera el digest diario unificado del proyecto: mails + reuniones + archivos."""
    # 1. Archivos modificados (cola de F1b)
    contenidos = []
    for item in items:
        ruta   = item["ruta"]
        nombre = item["nombre"]
        evento = item["evento"]
        if not Path(ruta).exists():
            log.warning(f"   Archivo no encontrado, ignorando: {nombre}")
            continue
        contenido = leer_contenido(ruta)
        if contenido and len(contenido.strip()) > 50:
            contenidos.append(f"--- ARCHIVO: {nombre} ({evento}) ---\n{contenido}")
    texto_archivos = "\n\n".join(contenidos) if contenidos else "Sin archivos modificados."

    # 2. Mails de ayer (los dejó Power Automate)
    texto_mails = leer_mails_del_dia(proyecto)

    # 3. Reuniones recientes (minutas generadas desde la última corrida)
    texto_reuniones = leer_reuniones_recientes(proyecto, cutoff_reuniones)

    # 4. Matriz de gobierno (riesgos + decisiones) — solo proyectos integrados.
    #    Es contexto para razonar/seguir; NO dispara el digest por sí sola.
    texto_matriz = leer_matriz(proyecto)

    # Decidir si hay algo que reportar
    hay_mails     = bool(texto_mails) and texto_mails.strip() not in ("", "[]")
    hay_reuniones = bool(texto_reuniones)
    hay_archivos  = bool(contenidos)
    if not (hay_mails or hay_reuniones or hay_archivos):
        log.info(f"[Digest] {proyecto}: sin novedades ayer, no se genera digest.")
        return

    log.info(f"[Digest] {proyecto} | mails: {hay_mails} | reuniones: {hay_reuniones} | archivos: {len(contenidos)}")

    contexto_actual = cargar_contexto(proyecto)

    # Histórico: últimos N digests diarios (solo contexto para detectar tendencias).
    texto_historico = leer_digests_recientes(proyecto, DIGESTS_HISTORICOS)

    prompt_path = PROMPTS_FOLDER / "prompt_digest_diario.txt"
    try:
        prompt_template = prompt_path.read_text(encoding="utf-8")
    except Exception as e:
        log.error(f"No se pudo leer prompt diario: {e}")
        return

    prompt = (prompt_template
              .replace("{{PROYECTO}}",   proyecto)
              .replace("{{FECHA}}",      date.today().strftime("%d/%m/%Y"))
              .replace("{{CONTEXTO}}",   contexto_actual)
              .replace("{{MATRIZ}}",     texto_matriz or "Sin matriz integrada para este proyecto.")
              .replace("{{MAILS}}",      texto_mails or "Sin mails.")
              .replace("{{REUNIONES}}",  texto_reuniones or "Sin reuniones.")
              .replace("{{ARCHIVOS}}",   texto_archivos)
              .replace("{{HISTORICO}}",  texto_historico or "Sin digests previos."))

    resultado = llamar_claude(prompt, modelo=MODELO_DIGEST, sin_pensar=True)
    texto_completo = resultado["texto"]

    # Separar digest del contexto actualizado
    partes = texto_completo.split("📌 CONTEXTO ACTUALIZADO DEL PROYECTO:")
    digest = partes[0].strip()

    if len(partes) > 1:
        nuevo_contexto = partes[1].strip()
        if nuevo_contexto and nuevo_contexto != "SIN_CAMBIOS":
            guardar_contexto(proyecto, nuevo_contexto)
            log.info(f"   Contexto actualizado para {proyecto}")

    escribir_novedad(proyecto, digest, "digest_diario")
    registrar_tokens(proyecto, "Digest_Diario", resultado["input_tokens"], resultado["output_tokens"],
                     modelo=MODELO_DIGEST)
    log.info(f"[Digest] Diario completado para {proyecto} (modelo: {MODELO_DIGEST})")


def digest_todos_los_proyectos():
    """Corre el digest diario unificado solo para proyectos pendientes."""
    if digest_ya_corrido_hoy():
        log.info("[Digest] Ya se corrió hoy para todos los proyectos, saltando.")
        return

    log.info("[Digest] Iniciando digest diario unificado...")
    cutoff_reuniones = _cutoff_reuniones()  # capturar antes de marcar proyectos
    cola = cargar_cola()
    ya_procesados = proyectos_procesados_hoy()

    hoy_inicio = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    for proyecto in PROYECTOS.values():
        if proyecto in ya_procesados:
            log.info(f"[Digest] {proyecto} ya procesado hoy, saltando.")
            continue

        items_a_procesar = []
        items_restantes = []

        for item in cola.get(proyecto, []):
            try:
                ts = datetime.fromisoformat(item["timestamp"])
                if ts < hoy_inicio:
                    items_a_procesar.append(item)
                else:
                    items_restantes.append(item)
            except Exception:
                items_restantes.append(item)

        try:
            digest_proyecto(proyecto, items_a_procesar, cutoff_reuniones)
            # Solo limpiar la cola y marcar si fue exitoso
            cola[proyecto] = items_restantes
            marcar_proyecto_procesado(proyecto)
        except Exception as e:
            log.error(f"[Digest] Error en {proyecto}: {e}")
            # No limpia cola ni marca, se reintenta en la próxima hora

    guardar_cola(cola)
    log.info("[Digest] Digest diario completado.")